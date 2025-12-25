import openpyxl
from io import BytesIO
from datetime import datetime
from app.models import Contact
from app import db
import os

def export_to_excel():
    """Экспорт контактов в Excel файл"""
    # Создаем новую рабочую книгу
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Телефонный справочник"
    
    # Заголовки
    headers = ['ФИО', 'Телефон рабочий', 'Телефон сотовый', 'Электронная почта', 'Примечания']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Данные
    contacts = Contact.query.all()
    for row, contact in enumerate(contacts, 2):
        ws.cell(row=row, column=1, value=contact.full_name)
        ws.cell(row=row, column=2, value=contact.work_phone)
        ws.cell(row=row, column=3, value=contact.mobile_phone)
        ws.cell(row=row, column=4, value=contact.email)
        ws.cell(row=row, column=5, value=contact.notes)
    
    # Сохраняем в буфер
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer

def validate_excel_file(file_path):
    """Проверяет, что файл действительно является Excel файлом"""
    # Проверяем расширение файла
    if not file_path.lower().endswith(('.xlsx', '.xls')):
        raise ValueError("Неверный формат файла. Требуется .xlsx или .xls")
    
    # Проверяем, что файл существует и доступен для чтения
    if not os.path.exists(file_path) or not os.access(file_path, os.R_OK):
        raise ValueError("Файл не существует или недоступен для чтения")
    
    # Проверяем размер файла (не более 16 МБ)
    if os.path.getsize(file_path) > 16 * 1024 * 1024:
        raise ValueError("Файл слишком большой. Максимальный размер 16 МБ")
    
    # Пытаемся открыть файл с помощью openpyxl для проверки его целостности
    try:
        wb = openpyxl.load_workbook(file_path)
        wb.active  # Проверяем, что лист доступен
    except Exception:
        raise ValueError("Файл поврежден или не является корректным Excel файлом")

def import_from_excel(file_path):
    """Импорт контактов из Excel файла"""
    # Проверяем безопасность файла перед импортом
    validate_excel_file(file_path)
    
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    # Пропускаем заголовки
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    
    imported_count = 0
    for row in rows:
        if row[0]:  # Проверяем, что ФИО не пустое
            # Очищаем данные от потенциально опасного содержимого
            full_name = str(row[0]).strip() if row[0] else ""
            work_phone = str(row[1]).strip() if row[1] else None
            mobile_phone = str(row[2]).strip() if row[2] else None
            email = str(row[3]).strip() if row[3] else None
            notes = str(row[4]).strip() if row[4] else None
            
            contact = Contact(
                full_name=full_name,
                work_phone=work_phone,
                mobile_phone=mobile_phone,
                email=email,
                notes=notes
            )
            db.session.add(contact)
            imported_count += 1
    
    db.session.commit()
    return imported_count